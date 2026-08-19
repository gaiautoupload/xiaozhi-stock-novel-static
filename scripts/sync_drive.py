#!/usr/bin/env python3
"""Sync the XiaoZhi publication package from Google Drive into same-origin static files.

Auth:
1) Try GDRIVE_SERVICE_ACCOUNT_JSON (recommended for private Drive files).
2) Fall back to public "anyone with link" downloads.

The service-account secret is used only inside GitHub Actions; it is never copied into web assets.
"""
from __future__ import annotations
from pathlib import Path
import os, json, re, io, hashlib, datetime, tempfile, shutil
import requests

ROOT=Path(__file__).resolve().parents[1]
DATA=ROOT/"data"
CONFIG=json.loads((ROOT/"config"/"site.json").read_text(encoding="utf-8"))
ERRORS=[]

def now_iso():
    return datetime.datetime.now(datetime.timezone.utc).astimezone().isoformat(timespec="seconds")

def authorized_session():
    raw=os.getenv("GDRIVE_SERVICE_ACCOUNT_JSON","").strip()
    if not raw: return None
    try:
        from google.oauth2 import service_account
        from google.auth.transport.requests import AuthorizedSession
        info=json.loads(raw)
        creds=service_account.Credentials.from_service_account_info(info,scopes=["https://www.googleapis.com/auth/drive.readonly"])
        return AuthorizedSession(creds)
    except Exception as e:
        ERRORS.append(f"service account init: {e}")
        return None

AUTH=authorized_session()
PUBLIC=requests.Session()
PUBLIC.headers["User-Agent"]="xiaozhi-static-sync/1.0"

def fetch_bytes(file_id:str)->bytes:
    if AUTH:
        u=f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media&supportsAllDrives=true"
        r=AUTH.get(u,timeout=45)
    else:
        u=f"https://drive.usercontent.google.com/download?id={file_id}&export=download&confirm=t"
        r=PUBLIC.get(u,timeout=45,allow_redirects=True)
    r.raise_for_status()
    b=r.content
    head=b[:300].lower()
    if b"<html" in head and (b"sign in" in head or b"google drive" in head):
        raise RuntimeError(f"Drive file {file_id} is not public and no working service account secret is configured.")
    return b

def fetch_json(file_id:str):
    return json.loads(fetch_bytes(file_id).decode("utf-8-sig"))

def atomic_write(path:Path,b:bytes):
    path.parent.mkdir(parents=True,exist_ok=True)
    tmp=path.with_suffix(path.suffix+".tmp")
    tmp.write_bytes(b); tmp.replace(path)

def write_json(path:Path,obj):
    atomic_write(path,json.dumps(obj,ensure_ascii=False,indent=2).encode("utf-8"))

def split_archive(text:str):
    matches=list(re.finditer(r'<a id="(U\d{8}-\d{4})"></a>',text))
    out={}
    for i,m in enumerate(matches):
        end=matches[i+1].start() if i+1<len(matches) else len(text)
        uid=m.group(1)
        chunk=text[m.end():end].strip()
        # Remove archive-specific heading while retaining frontmatter/body.
        chunk=re.sub(r'^##\s+U\d{8}-\d{4}｜.*?\n+','',chunk,count=1)
        out[uid]=chunk.strip()+"\n"
    return out

def parse_frontmatter_chars(md:str):
    m=re.match(r'^\s*---\s*\n(.*?)\n---',md,re.S)
    if not m:return []
    block=m.group(1)
    # supports YAML list form
    cm=re.search(r'(?m)^characters:\s*(.*)$',block)
    if not cm:return []
    tail=cm.group(1).strip()
    if tail.startswith("["):
        return re.findall(r'["\']([^"\']+)["\']',tail)
    after=block[cm.end():]
    out=[]
    for line in after.splitlines():
        if re.match(r'^\w[\w_-]*:',line): break
        lm=re.match(r'^\s*-\s*(.+?)\s*$',line)
        if lm: out.append(lm.group(1).strip(" \"'"))
    return out

def build_catalog(feed):
    entries=[]
    for u in feed.get("updates",[]):
        uid=u.get("update_id")
        cp=DATA/"chapters"/f"{uid}.md"
        sp=DATA/"stocks"/f"{uid}.stocks.json"
        vp=DATA/"visuals"/f"{uid}.market.svg"
        entries.append({
            "sequence":u.get("sequence"),"update_id":uid,"published_at":u.get("published_at"),
            "title":u.get("title") or uid,"chapter_path":str(cp.relative_to(ROOT)).replace("\\","/") if cp.exists() else None,
            "stocks_path":str(sp.relative_to(ROOT)).replace("\\","/") if sp.exists() else None,
            "market_svg_path":str(vp.relative_to(ROOT)).replace("\\","/") if vp.exists() else None,
            "backfill":bool(u.get("backfill",False))
        })
    return {"generated_at":now_iso(),"latest_update_id":feed.get("latest_update_id"),"sequence":feed.get("sequence"),"entries":entries}

RATING={"A+":10,"A":9,"A-":8,"B+":7,"B":6,"B-":5,"C+":4,"C":3,"C-":2,"D":1}

def build_stock_history(catalog):
    snapshots=[]
    for e in sorted(catalog["entries"],key=lambda x:x.get("sequence") or 0):
        if not e.get("stocks_path"): continue
        try:
            d=json.loads((ROOT/e["stocks_path"]).read_text(encoding="utf-8"))
            snapshots.append((e,d))
        except Exception as ex: ERRORS.append(f"stocks parse {e['update_id']}: {ex}")
    events=[]; prev={}
    for e,d in snapshots:
        current={str(x.get("ticker")):x for x in d.get("watchlist",[]) if x.get("ticker")}
        explicit=d.get("selection_changes",[])
        explicit_by_ticker={str(x.get("ticker")):x for x in explicit if x.get("ticker")}
        for t,s in current.items():
            p=prev.get(t); x=explicit_by_ticker.get(t)
            if x:
                action=x.get("action","變更")
                reason=x.get("reason") or x.get("why") or "原始快照有變更事件，但未提供文字理由。"
                src="explicit"
            elif p is None:
                action="新增"
                reason="；".join((s.get("evidence") or [])[:2]) or "首次出現在可取得的結構化觀察快照。"
                src="auto_diff"
            else:
                pr,cr=p.get("rating"),s.get("rating")
                if RATING.get(cr,0)>RATING.get(pr,0): action="升評"
                elif RATING.get(cr,0)<RATING.get(pr,0): action="降評"
                elif p.get("positioning")!=s.get("positioning"): action="定位調整"
                else: action="維持"
                newris=[z for z in s.get("risks",[]) if z not in p.get("risks",[])]
                newev=[z for z in s.get("evidence",[]) if z not in p.get("evidence",[])]
                basis=(newris if action=="降評" else newev) or (s.get("confirm_signals") or [])[:1]
                reason="；".join(basis[:2]) if basis else "前後快照未提供獨立變更註記；評級／定位依原始快照維持。"
                src="auto_diff"
            events.append({
                "sequence":e.get("sequence"),"update_id":e.get("update_id"),"as_of":d.get("as_of") or e.get("published_at"),
                "ticker":t,"name":s.get("name"),"action":action,"previous_rating":p.get("rating") if p else None,
                "rating":s.get("rating"),"positioning":s.get("positioning"),"reason":reason,"reason_source":src,
                "evidence":s.get("evidence",[]),"risks":s.get("risks",[]),"confirm_signals":s.get("confirm_signals",[]),
                "invalidations":s.get("invalidations",[]),"exit_logic":s.get("exit_logic"),"trade_status":s.get("trade_status")
            })
        # Detect removals and make inference status explicit.
        for t,p in prev.items():
            if t in current: continue
            x=explicit_by_ticker.get(t)
            reason=(x or {}).get("reason") if x else None
            if not reason:
                inv="；".join((p.get("invalidations") or [])[:2])
                reason=f"本次觀察清單不再包含；原始快照未提供明確汰換理由。最後已知失效條件：{inv or '未提供'}"
            events.append({
                "sequence":e.get("sequence"),"update_id":e.get("update_id"),"as_of":d.get("as_of") or e.get("published_at"),
                "ticker":t,"name":p.get("name"),"action":"汰換","previous_rating":p.get("rating"),"rating":"OUT",
                "positioning":p.get("positioning"),"reason":reason,"reason_source":"explicit" if x else "auto_diff",
                "evidence":p.get("evidence",[]),"risks":p.get("risks",[]),"confirm_signals":p.get("confirm_signals",[]),
                "invalidations":p.get("invalidations",[]),"exit_logic":p.get("exit_logic"),"trade_status":"OUT"
            })
        prev=current
    return {"generated_at":now_iso(),"events":events,"current":list(prev.values()),"snapshot_count":len(snapshots)}

def build_character_index(catalog):
    chars={}
    for e in catalog["entries"]:
        if not e.get("chapter_path"): continue
        try: md=(ROOT/e["chapter_path"]).read_text(encoding="utf-8")
        except: continue
        for name in parse_frontmatter_chars(md):
            chars.setdefault(name,[]).append({"sequence":e["sequence"],"update_id":e["update_id"],"title":e["title"]})
    return {"generated_at":now_iso(),"characters":[{"name":k,"appearances":v} for k,v in sorted(chars.items(),key=lambda kv:(-len(kv[1]),kv[0]))]}

def sync_sheet():
    sid=CONFIG.get("sourceSheetFileId")
    if not sid:return
    try:
        if AUTH:
            mt="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            r=AUTH.get(f"https://www.googleapis.com/drive/v3/files/{sid}/export",params={"mimeType":mt},timeout=60);r.raise_for_status();blob=r.content
        else:
            r=PUBLIC.get(f"https://docs.google.com/spreadsheets/d/{sid}/export?format=xlsx",timeout=60);r.raise_for_status();blob=r.content
        from openpyxl import load_workbook
        wb=load_workbook(io.BytesIO(blob),read_only=True,data_only=True)
        sheets={}
        for ws in wb.worksheets:
            vals=list(ws.iter_rows(values_only=True))
            if not vals: sheets[ws.title]=[];continue
            headers=[str(x).strip() if x is not None and str(x).strip() else f"col_{i+1}" for i,x in enumerate(vals[0])]
            rows=[]
            for row in vals[1:]:
                if not any(x is not None and str(x)!="" for x in row):continue
                rows.append({headers[i]:row[i] if i<len(row) else None for i in range(len(headers))})
            sheets[ws.title]=rows
        write_json(DATA/"sheet-tabs.json",{"generated_at":now_iso(),"sheets":sheets})
    except Exception as e:
        ERRORS.append(f"sheet export skipped: {e}")

def main():
    DATA.mkdir(exist_ok=True)
    # Fetch indices first. If these fail, abort without deleting prior snapshots.
    manifest=fetch_json(CONFIG["manifestFileId"])
    feed=fetch_json(CONFIG["feedIndexFileId"])
    market_index=fetch_json(CONFIG["marketIndexFileId"])
    write_json(DATA/"manifest.json",manifest);write_json(DATA/"feed.index.json",feed);write_json(DATA/"market.index.json",market_index)

    # Hydrate archive chapter sources.
    for archive_path,fid in CONFIG.get("archiveFileIds",{}).items():
        if not archive_path.startswith("chapters/"):continue
        try:
            text=fetch_bytes(fid).decode("utf-8-sig")
            atomic_write(DATA/"archives"/Path(archive_path).name,text.encode("utf-8"))
            for uid,chunk in split_archive(text).items():
                atomic_write(DATA/"chapters"/f"{uid}.md",chunk.encode("utf-8"))
        except Exception as e: ERRORS.append(f"archive {archive_path}: {e}")

    # Hydrate individually-addressed release files.
    for u in feed.get("updates",[]):
        uid=u.get("update_id")
        for field,subdir,suffix in [
            ("chapter_file_id","chapters",".md"),("stocks_file_id","stocks",".stocks.json"),("market_svg_file_id","visuals",".market.svg")
        ]:
            fid=u.get(field)
            if not fid:continue
            path=DATA/subdir/f"{uid}{suffix}"
            try: atomic_write(path,fetch_bytes(fid))
            except Exception as e: ERRORS.append(f"{uid} {field}: {e}")

    # Hydrate market-intel history.
    for m in market_index.get("updates",[]):
        mid=m.get("market_id")
        for field,suffix in [("json_file_id",".market.json"),("svg_file_id",".market.svg")]:
            fid=m.get(field)
            if not fid:continue
            path=DATA/"market-intel"/f"{mid}{suffix}"
            try: atomic_write(path,fetch_bytes(fid))
            except Exception as e: ERRORS.append(f"{mid} {field}: {e}")

    catalog=build_catalog(feed);write_json(DATA/"catalog.json",catalog)
    write_json(DATA/"stock-history.json",build_stock_history(catalog))
    write_json(DATA/"character-index.json",build_character_index(catalog))
    sync_sheet()
    write_json(DATA/"sync-status.json",{
        "last_success":now_iso(),"mode":"service_account" if AUTH else "public_drive",
        "feed_sequence":feed.get("sequence"),"market_sequence":market_index.get("sequence"),
        "chapter_count":sum(1 for e in catalog["entries"] if e["chapter_path"]),
        "stocks_snapshot_count":sum(1 for e in catalog["entries"] if e["stocks_path"]),
        "errors":ERRORS,"message":"Drive 同步完成；有警告時保留舊快照，不刪除既有資料。"
    })
    print(json.dumps({"feed":feed.get("sequence"),"market":market_index.get("sequence"),"errors":len(ERRORS)},ensure_ascii=False))
if __name__=="__main__": main()
